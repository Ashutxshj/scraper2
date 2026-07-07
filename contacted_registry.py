"""Persistent "already contacted" registry — contacted_businesses.xlsx.

Every business that went out in a successfully-sent Resend report is recorded
here. Future runs load the registry and drop those businesses before building
the report, so the same lead is never emailed twice.

A business matches the registry if ANY of its identity keys matches:
  * place_id (strongest — stable Google Maps id)
  * normalized phone number (last 10 digits)
  * business name + address (lowercased)
Multiple keys per business catch the same place resurfacing from a different
source (e.g. Apify one run, seed CSV the next) without a place_id.
"""

import os
import re
from datetime import datetime

import config

COLUMNS = [
    "Business Name", "Phone Number", "Email", "Address",
    "Category", "Lead Type", "Place ID", "First Emailed At",
]


def _norm_phone(phone: str) -> str:
    """Digits only, last 10 — collapses '+91 98110 12345' / '09811012345' etc."""
    digits = re.sub(r"\D", "", phone or "")
    return digits[-10:] if len(digits) >= 10 else digits


def identity_keys(name: str, phone: str, address: str, place_id: str) -> set[str]:
    keys = set()
    if place_id:
        keys.add(f"id:{place_id}")
    norm = _norm_phone(phone)
    if norm:
        keys.add(f"ph:{norm}")
    name = (name or "").strip().lower()
    if name:
        keys.add(f"na:{name}|{(address or '').strip().lower()}")
    return keys


def target_keys(target: dict) -> set[str]:
    """Identity keys for a sourcer target dict."""
    return identity_keys(target.get("business_name", ""), target.get("phone", ""),
                         target.get("address", ""), target.get("place_id", ""))


def _row_keys(row: dict) -> set[str]:
    """Identity keys for a report/registry row (report-column names)."""
    return identity_keys(row.get("Business Name", ""), str(row.get("Phone Number", "")),
                         row.get("Address", ""), str(row.get("Place ID", "")))


def load_keys() -> set[str]:
    """All identity keys of already-contacted businesses ({} if no registry yet)."""
    path = config.CONTACTED_FILE
    if not os.path.exists(path):
        return set()
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(f"[registry] WARN: openpyxl not installed — cannot read {path}; "
              "already-contacted filtering is OFF (pip install openpyxl)")
        return set()

    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    keys: set[str] = set()
    count = 0
    if header:
        idx = {str(h): i for i, h in enumerate(header) if h is not None}
        for values in rows:
            row = {col: (values[i] if i < len(values) and values[i] is not None else "")
                   for col, i in idx.items()}
            got = _row_keys(row)
            if got:
                keys |= got
                count += 1
    wb.close()
    print(f"[registry] {count} previously-contacted businesses loaded from {path}")
    return keys


def record(rows: list[dict]) -> int:
    """Append report rows to the registry workbook. Returns how many were added.
    Never raises — a registry failure must not undo a successful send."""
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[registry] WARN: openpyxl not installed — businesses NOT recorded; "
              "they will reappear in the next run (pip install openpyxl)")
        return 0

    path = config.CONTACTED_FILE
    try:
        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb.active
            existing = set()
            header = [c.value for c in ws[1]]
            idx = {str(h): i for i, h in enumerate(header) if h is not None}
            for values in ws.iter_rows(min_row=2, values_only=True):
                existing |= _row_keys({col: (values[i] if i < len(values) and
                                             values[i] is not None else "")
                                       for col, i in idx.items()})
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Contacted"
            ws.append(COLUMNS)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for i, col in enumerate(COLUMNS, 1):
                ws.column_dimensions[get_column_letter(i)].width = len(col) + 12
            ws.freeze_panes = "A2"
            existing = set()

        stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        added = 0
        for row in rows:
            if _row_keys(row) & existing:
                continue  # e.g. two branches sharing one phone number this run
            existing |= _row_keys(row)
            ws.append([
                row.get("Business Name", ""), row.get("Phone Number", ""),
                row.get("Email", ""), row.get("Address", ""),
                row.get("Category", ""), row.get("Lead Type", ""),
                row.get("Place ID", ""), stamp,
            ])
            added += 1
        wb.save(path)
        return added
    except Exception as exc:  # registry write must never break the pipeline
        print(f"[registry] WARN: could not update {path}: {exc}")
        return 0
