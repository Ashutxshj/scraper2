"""Orchestrator: pick category -> source Google Maps places with NO website ->
STRICT VALIDATION GATE -> classify Goldenrod / New Bark -> CSV + styled XLSX -> Resend.

Lead tiers (every lead has no website by construction):
  * Goldenrod — rating >= 4.5 with 500+ reviews. Rendered with a golden-yellow
    row background in the XLSX.
  * New Bark  — every other no-website business.

Usage:
  python main.py                 # full run (Apify / Places API / seed_places.csv)
  python main.py --mock          # built-in demo targets, no API keys needed
  python main.py --limit 25      # cap number of sourced targets
  python main.py --no-email      # build the report but skip Resend delivery
  python main.py --category 3    # preselect a business category (skips the menu)

At startup a numbered menu of BUSINESS_CATEGORIES is shown — enter a number to
source only that category (e.g. 1 = dentist), or press Enter / 0 for all.
"""

import argparse
import csv
import os
import shutil
import sys
from datetime import datetime

import config
import contacted_registry
from delivery import send_report
from target_sourcer import source_targets


def pick_category(preselected: int | None) -> str | None:
    """Numbered category menu. Returns the chosen category, or None for all."""
    categories = config.BUSINESS_CATEGORIES
    if preselected is not None:
        if preselected == 0:
            print("[main] --category 0 -> ALL categories")
            return None
        if 1 <= preselected <= len(categories):
            print(f"[main] --category {preselected} -> '{categories[preselected - 1]}'")
            return categories[preselected - 1]
        print(f"[main] --category {preselected} out of range — using ALL categories")
        return None

    print("\nWhich business category should this run target?")
    for i, cat in enumerate(categories, 1):
        print(f"  {i:2d}. {cat}")
    print("   0. all categories")
    while True:
        try:
            raw = input("Enter a number [0]: ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\n[main] no selection — using ALL categories")
            return None
        if raw in ("", "0"):
            return None
        if raw.isdigit() and 1 <= int(raw) <= len(categories):
            choice = categories[int(raw) - 1]
            print(f"[main] targeting category: {choice}")
            return choice
        print(f"Invalid choice '{raw}' — enter 0-{len(categories)}")


def classify(target: dict) -> str:
    """Goldenrod iff rating >= 4.5 AND 500+ reviews; everything else is New Bark."""
    rating = target.get("rating")
    reviews = target.get("reviews") or 0
    if (rating is not None
            and rating >= config.GOLDENROD_MIN_RATING
            and reviews >= config.GOLDENROD_MIN_REVIEWS):
        return config.LEAD_TYPE_GOLDENROD
    return config.LEAD_TYPE_NEW_BARK


def write_xlsx(rows: list[dict], path: str) -> str | None:
    """Styled workbook: bold header, golden-yellow fill on Goldenrod rows.
    Returns the path, or None if openpyxl isn't installed (CSV still exists)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[main] openpyxl not installed — skipping styled XLSX "
              "(pip install openpyxl)")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(config.CSV_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    gold = PatternFill(fill_type="solid", fgColor=config.GOLDENROD_FILL_HEX)
    for row in rows:
        ws.append([row[col] for col in config.CSV_COLUMNS])
        if row["Lead Type"] == config.LEAD_TYPE_GOLDENROD:
            for cell in ws[ws.max_row]:
                cell.fill = gold

    for i, col in enumerate(config.CSV_COLUMNS, 1):
        width = max([len(col)] + [len(str(r[col])) for r in rows])
        ws.column_dimensions[get_column_letter(i)].width = min(width + 2, 50)

    ws.freeze_panes = "A2"
    wb.save(path)
    return path


def run(mock: bool, limit: int, no_email: bool, category: str | None = None) -> int:
    # --- Stage 1: source no-website places from Google Maps ---
    targets = source_targets(mock=mock, category=category)

    # Already-emailed businesses are dropped BEFORE --limit so the cap fills
    # with fresh leads. Mock runs skip the registry so smoke tests stay green.
    if not mock:
        contacted = contacted_registry.load_keys()
        if contacted:
            before = len(targets)
            targets = [t for t in targets
                       if not (contacted_registry.target_keys(t) & contacted)]
            if before - len(targets):
                print(f"[main] {before - len(targets)} already-contacted "
                      f"businesses skipped (see {config.CONTACTED_FILE})")

    if limit > 0:
        targets = targets[:limit]
    if not targets:
        print("[main] no targets sourced — nothing to do")
        return 1

    rows: list[dict] = []
    stats = {"sourced": len(targets), "no_contact": 0,
             config.LEAD_TYPE_GOLDENROD: 0, config.LEAD_TYPE_NEW_BARK: 0}

    for i, target in enumerate(targets, 1):
        name = target["business_name"]
        phone, email = target.get("phone") or "", target.get("email") or ""

        # --- Stage 2: STRICT VALIDATION GATE ---
        # Rule: if email IS NULL and phone IS NULL, discard the row entirely.
        if not email and not phone:
            print(f"[main] ({i}/{len(targets)}) {name}: no email AND no phone — discarded")
            stats["no_contact"] += 1
            continue

        # --- Stage 3: tier classification ---
        lead_type = classify(target)
        stats[lead_type] += 1
        rating = target.get("rating")
        print(f"[main] ({i}/{len(targets)}) {name}: {lead_type} "
              f"(rating={rating if rating is not None else 'n/a'}, "
              f"reviews={target.get('reviews') or 0}, phone={phone or '-'})")

        rows.append({
            "Business Name": name,
            "Category": target.get("category") or "",
            "Phone Number": phone,
            "Email": email,
            "Rating": rating if rating is not None else "",
            "Reviews": target.get("reviews") or 0,
            "Address": target.get("address") or "",
            "Lead Type": lead_type,
            # Not a report column — carried for the contacted registry only.
            "Place ID": target.get("place_id") or "",
        })

    # Goldenrod first, then by review count within each tier.
    rows.sort(key=lambda r: (r["Lead Type"] != config.LEAD_TYPE_GOLDENROD,
                             -(r["Reviews"] or 0)))

    # --- Stage 4: write CSV + styled XLSX to the archive folder ---
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.join(config.OUTPUT_DIR, f"{config.OUTPUT_BASENAME}_{stamp}")
    csv_path, xlsx_path = f"{base}.csv", f"{base}.xlsx"
    latest_path = os.path.join(config.OUTPUT_DIR, f"{config.OUTPUT_BASENAME}_latest.csv")

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=config.CSV_COLUMNS,
                                extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    shutil.copyfile(csv_path, latest_path)  # stable path for quick access
    xlsx_written = write_xlsx(rows, xlsx_path)

    goldenrod = stats[config.LEAD_TYPE_GOLDENROD]
    new_bark = stats[config.LEAD_TYPE_NEW_BARK]
    print(
        f"\n[main] done: sourced={stats['sourced']} no_contact={stats['no_contact']} "
        f"kept={len(rows)} (Goldenrod={goldenrod}, New Bark={new_bark})"
    )
    print(f"[main] CSV archived -> {csv_path}")
    print(f"[main] latest copy  -> {latest_path}")
    if xlsx_written:
        print(f"[main] styled XLSX  -> {xlsx_written}")

    # --- Stage 5: delivery (wa1.txt theory) ---
    if no_email:
        print("[main] --no-email set — skipping delivery")
    elif not rows:
        print("[main] 0 validated leads — skipping delivery")
    else:
        # One attachment only: the styled XLSX (all leads, both tiers, Goldenrod
        # rows highlighted). CSV is the fallback if openpyxl isn't installed.
        sent = send_report([xlsx_written or csv_path],
                           goldenrod=goldenrod, new_bark=new_bark)
        # Record ONLY what actually went out: an unsent lead stays eligible
        # for the next run. Mock leads never enter the registry.
        if sent and not mock:
            added = contacted_registry.record(rows)
            print(f"[main] {added} businesses recorded in {config.CONTACTED_FILE} "
                  "— they will be skipped in future runs")

    return 0


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Delhi NCR no-website business leads pipeline")
    parser.add_argument("--mock", action="store_true", help="use built-in demo targets")
    parser.add_argument("--limit", type=int, default=0, help="cap sourced targets")
    parser.add_argument("--no-email", action="store_true", help="skip Resend delivery")
    parser.add_argument("--category", type=int, default=None,
                        help="business category number (1-based, see startup menu); "
                             "omit for the interactive menu, 0 for all")
    args = parser.parse_args()
    chosen = pick_category(args.category)
    sys.exit(run(mock=args.mock, limit=args.limit, no_email=args.no_email,
                 category=chosen))
